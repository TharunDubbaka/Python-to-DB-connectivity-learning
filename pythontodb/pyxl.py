import openpyxl

wb=openpyxl.load_workbook("c:\\Users\\dubba\\Downloads\\PythonXL.xlsx")
sheet=wb.active
for i in sheet.iter_rows(values_only=True):
    print(i)
newvals=[["Mouse","h8",93],['keyboard',"h7",94],['HI',"g8",87]]
for i in newvals:
    sheet.append(i)
print("After appending \n")
for i in sheet.iter_rows(values_only=True):
    print(i)
sheet.delete_rows(idx=4)
print("After deleting \n")
for i in sheet.iter_rows(values_only=True):
    print(i)
wb.save("PythonXL.xlsx")